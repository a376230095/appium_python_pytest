from openpyxl import load_workbook
#读取excel表格
wb = load_workbook(filename = 'empty_book.xlsx')
#读取哪一个sheet
sheet_ranges = wb['range names']
#打印出sheet中的值
print(sheet_ranges['D18'].value)

for i in range(1,31):
    #读取每一个cell的值
    print(sheet_ranges.cell(column=1,row=i).value)