from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#创建一个execl的工作本
wb = Workbook()
#给execl命名
dest_filename = 'empty_book1.xlsx'
#激活execl表格
ws1 = wb.active
#给第一个sheet命名
ws1.title = "range names"
#获取1到第40行
for row in range(1, 40):
    #每1行增加600个列数据
     ws1.append(range(600))
#创建另一个sheet，标题为Pi
ws2 = wb.create_sheet(title="Pi")
#直接给F5赋值3.14
ws2['F5'] = 3.14
#创建另一个sheet，标题为Date
ws3 = wb.create_sheet(title="Data")
#取第10行和第19行
for row in range(10, 20):
    #取第26列和第54列
     for col in range(27, 54):
         #这个_应该是固定写法吧
         #给一个cell赋值，定义了colume列和row行，还有value的值
         ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
#打印出AA10的值出来
print(ws3['AA10'].value)
#保存execl表格
wb.save(filename = dest_filename)